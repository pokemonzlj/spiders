# -*- coding: utf-8 -*-
from web_control import web_test
# import requests
from requests_html import HTMLSession
import json
import re
import sys
import os
import time
import threading
import queue
from bs4 import BeautifulSoup
import openpyxl

output_print = web_test.output_print
class do_test:
    def __init__(self):
        self.total_item_list={}  #用来保存完整的商品地址,键是商品地址，值是商品的来源目录
        self.total_jpg_list=[]  #用来保存完整的图片链接地址
        headers= {
                  "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.212 Safari/537.36",
                       "content-type": "application/x-www-form-urlencoded; charset=UTF-8"}#"User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Safari/537.36",
        self.sessions = HTMLSession()
        self.sessions.headers = headers
        self.queue = queue.Queue(100) #用Queue构造一个大小为100的线程安全的先进先出队列
        self.item_queue = queue.Queue(200)  #将商品地址存入队列
        self.category_list = []  #用来保存类目页地址
        # self.item_list=[]  #用来保存商品地址
        self.sheet = ''
        self.father_url = ''  #用来保存来源地址
        self.start_row = 2  #起始写入表格的行

    def create_excel(self):
        '''根据执行时间新建excel'''
        excel_new=openpyxl.Workbook()
        sheet1=excel_new.active
        sheet1.title="analyse_result"
        sheet1.cell(row=1, column=1).value = "来源页面"
        sheet1.cell(row=1, column=2).value = "请求URL"
        sheet1.cell(row=1, column=3).value = "状态码"
        sheet1.cell(row=1, column=4).value = "canonical"
        sheet2=excel_new.create_sheet('items')
        sheet2.cell(row=1, column=1).value = "来源页面"
        sheet2.cell(row=1, column=2).value = "商品链接"
        t = time.strftime("%Y_%m_%d_%H_%M", time.localtime())
        filename = 'results/'+ t + '_result.xlsx'
        excel_new.save(filename)
        return filename

    def read_excel(self, excel_name='', sheetnum=1):
        '''预读取excel，方便后面写入'''
        self.wb = openpyxl.load_workbook(excel_name)
        sheetname = self.wb.sheetnames
        print("Open excel %s sheet:%s" %(excel_name, sheetnum))
        self.sheet = self.wb.get_sheet_by_name(sheetname[sheetnum-1])

    def save_excel(self, excel_name=''):
        '''保存excel'''
        self.wb.save(excel_name)

    def get_picture_address(self, add='', relative=False):
        '''商详页获取.jpg格式的地址,relative参数控制，为True则获取相对路径'''
        page_result=self.sessions.get(add)
        status = page_result.status_code
        print("Item address:%s  Item Status_code:%s" % (add,status))
        if status != 200:
            return False
        # match="((https:\/\/).*?(\.jpg))"
        # match = "(https:\/\/.*?\.jpg)"
        match = '(https:\/\/[a-zA-Z0-9\/\.\_\\~-]*?\.jpg)'
        if relative:
            match = "(?<=:\")\/.*?(\.jpg)"
        jpg_link=re.findall(match, page_result.text)
        total_jpg_count=len(jpg_link)
        if relative:
            print("Current page have %s relative jpg link." % total_jpg_count)
        else:
            print("Current page have %s jpg link."%total_jpg_count)
        if total_jpg_count > 0:
            for jpg_add in jpg_link:
                if jpg_add not in self.total_jpg_list:
                    self.total_jpg_list.append(jpg_add)
                # status = requests.get(jpg_add).status_code
                # status=str(status)
                # print(jpg_add+'  '+status)
                    print(jpg_add)
            return jpg_link
        return False

    def get_category_list(self, add='https://www.aosom.fr', get_all_address=False):
        '''主页上获取相对+绝对地址，默认只获取相对地址'''
        page_result = self.sessions.get(add)
        # category_list=page_result.html.links
        # return True
        # match = "(?<=href=\")(?!https)[a-zA-Z0-9\/~-]+?\/(?=\")"   #适用于M2的类目
        match = "((?<=href=(\"|\'))(?!https)[a-zA-Z0-9\/\.~-]+?(?=\"|\'))"
        category_list=re.findall(match, page_result.text)
        total_category_count = len(category_list)
        # if total_category_count <10:
        #     # match = "(?<=href=(\"|\'))(?!https)[a-zA-Z0-9\/~-]*?\.html(?=(\"|\'))"
        #     match = "(?<=href=(\"|\'))(?!https)[a-zA-Z0-9\/~-]*?(\.html|\/)(?=(\"|\'))"
        #     category_list = re.findall(match, page_result.text)
        #     total_category_count = len(category_list)
        print("Current page get %s relative links." % total_category_count)
        if get_all_address:
            match_absolute = "https:\/\/[a-zA-Z0-9\/\.~\-\_]*(?=\"|\')"  #https开头的绝对路径
            absolute_link_list = re.findall(match_absolute, page_result.text)
            total_absolute_link_count = len(absolute_link_list)
            print("Current page get %s absolute links." % total_absolute_link_count)
        real_category_list=[]
        for category_list_add in category_list:
            # category_list_add=add+category_list_add
            # print(category_list_add[0])
            category_list_add=list(category_list_add)
            # category_list_add[0]=category_list_add[0].lower()   #大小写不同算不同的地址，所以无需大小写转化
            if category_list_add[0] not in real_category_list:
                real_category_list.append(category_list_add[0])
        if get_all_address:
            for absolute_link in absolute_link_list:
                if absolute_link not in real_category_list:
                    real_category_list.append(absolute_link)
        real_category_count = len(real_category_list)
        print("Current page actually have %s different links." % real_category_count)
        for i in range(real_category_count):
            if real_category_list[i].startswith('http'):
                continue
            elif real_category_list[i].startswith('/'):
                real_category_list[i]=add+real_category_list[i]   #对地址做统一的拼接
            else:
                real_category_list[i] = add + '/' + real_category_list[i]
            # print(real_category_list[i])
            # real_category_list[i]=real_category_list[i].strip('/')   #在写入的时候就去掉地址末尾的/,这个会造成301，所以原先是啥样就保持啥样
            self.queue.put(real_category_list[i])
        # return real_category_list

    def get_item_list(self, add=''):
        '''获取类目/搜索页面上商品链接,相对的地址  m2:href="/item/aosom-kids-12v-rc-2-seater-ride-on-police-truck-led-lights-mp3-white~370-082WT.html"
        会出异常的链接格式往往不带/item/，所以需要区分来判断'''
        page_result = self.sessions.get(add)
        match = "(?<=href=\")\/item(?!https).*?\.html(?=\")"
        item_list=re.findall(match, page_result.text)
        total_item_count = len(item_list)
        print("Current page get %s item link." % total_item_count)
        # for item_list_add in item_list:
        return item_list

    def get_status(self, add='', need_print_200=False):
        '''获取页面状态码,并返回状态码，将符合类目格式的地址存入数组'''
        page_result = self.sessions.get(add, allow_redirects=False)
        status = page_result.status_code
        if need_print_200:
            self.sheet.cell(row=self.start_row, column=1).value = self.father_url
            self.sheet.cell(row=self.start_row, column=2).value = add
            self.sheet.cell(row=self.start_row, column=3).value = status
            self.start_row += 1
            print("URL add:%s  URL Status_code:%s" % (add, status))
        elif status != 200:
            self.sheet.cell(row=self.start_row, column=1).value = self.father_url
            self.sheet.cell(row=self.start_row, column=2).value = add
            self.sheet.cell(row=self.start_row, column=3).value = status
            self.start_row += 1
            print("URL add:%s  URL Status_code:%s" % (add, status))
            return False
        if add.endswith(".png") or add.endswith(".ico") or add.endswith(".jpg") or add.endswith(".css") or add.endswith(".woff2") or add.endswith(".gif") or add.endswith(".svg"):
            return False
        elif '/page/' in add or '/item/' in add or '/account/' in add:  #把活动页也算进类目页中or '/activity/' in add
            return False
        elif len(add)<22:  #字符长度非常短的一定就是主页地址
            return False
        self.category_list.append(add)

    def get_status_total(self, need_print_200=False):
        '''获取页面状态码,并返回状态码和页面内容'''
        while not self.queue.empty():
            add = self.queue.get()
            self.get_status(add, need_print_200)
            self.match_canonical(add)
            # return status  #循环判断为true，所以不能return

    def get_total_count(self, add='', webtype='m2'):
        '''获取类目或搜索页得到商品的个数，<span id="total_count">34</span>
        pop:<div class="result-title">Showing 1-21 of 59 Results</div>  Weergeven 1-21 of 110 Resultaten
        pop:更新 直接获取<div class="savePMes" pages="16" pageSize="21" total="330" pageNum="1"></div>'''
        # status = page_result.status_code
        # print("URL add:%s  URL Status_code:%s" % (add, status))
        # if status != 200:
        #     return False
        # if add.endswith(".png") or add.endswith(".ico") or add.endswith(".jpg") or add.endswith(".css"):
        #     return False
        # elif '/page/' in add or '/activity/' in add or '/item/' in add or '/account/' in add:
        #     return False
        page_result = self.sessions.get(add, allow_redirects=False)
        if webtype=='pop':
            # page_result.html.render(timeout=20)   #页面有的特别大默认的8秒加载不完，把延迟改大
            # results = page_result.html.xpath("//div[@class='result-title']/text()")
            soup = BeautifulSoup(page_result.text, 'lxml')
            results=soup.find(attrs={"class": "savePMes"}) #name='div',
            # print(results)
            if results:
                result= results.attrs  #参数返回字典形式
                return result['total']
            return 0
        match = "(?<=<span id=\"total_count\">).*?(?=<\/span>)"
        total_count = re.findall(match, page_result.text)
        if len(total_count)==0:
            return 0
        return total_count[0]

    def get_total_item_list(self, add='', webtype='m2'):
        '''从目录获取其下所有的商品链接'''
        if self.category_list:
            for category in self.category_list:  #经过前面的筛选，category一定是200的链接
                category_item_total_count = self.get_total_count(category, webtype)
                if category_item_total_count:
                    print("This category totally get %s item count." % category_item_total_count)
                category_item_total_count = int(category_item_total_count)
                if webtype == 'pop':
                    if 0 < category_item_total_count <= 24:
                        item_list = self.get_item_list(category)
                        for item in item_list:
                            item = add + item
                            if item not in self.total_item_list.keys():
                                self.total_item_list[item]=category   #把没统计的地址按地址：来源地址的格式写入字典
                                # print("item address:%s"%item)
                                # self.get_picture_address(item, True)
                    elif category_item_total_count > 24:
                        get_time = category_item_total_count // 24 + 1
                        for i in range(get_time):
                            item_list = self.get_item_list(category + "?page=%s" % (i + 1))
                            for item in item_list:
                                item = add + item
                                if item not in self.total_item_list.keys():
                                    self.total_item_list[item] = category
                                    # print("item address:%s"%item)
                                    # self.get_picture_address(item, True)
                    continue
                if 0 < category_item_total_count <= 100:
                    item_list = self.get_item_list(category + "?product_list_limit=100")
                    for item in item_list:
                        item = add + item
                        if item not in self.total_item_list.keys():
                            self.total_item_list[item] = category  # 把没统计的地址按地址：来源地址的格式写入字典
                            # print("item address:%s"%item)
                            # self.get_picture_address(item, True)
                elif category_item_total_count > 100:
                    get_time = category_item_total_count // 100 + 1
                    for i in range(get_time):
                        item_list = self.get_item_list(category + "?product_list_limit=100&p=%s" % (i + 1))
                        for item in item_list:
                            item = add + item
                            if item not in self.total_item_list.keys():
                                self.total_item_list[item] = category  # 把没统计的地址按地址：来源地址的格式写入字典
                                # print("item address:%s"%item)
                                # self.get_picture_address(item, True)

    def write_item_list(self, excel_name=''):
        '''把字典对应的地址倒过来写入excel'''
        self.read_excel(excel_name, 2)
        if self.total_item_list:
            start_row=2
            # self.sheet.cell(row=start_row, column=6).value ="item url"
            print("Write item url into sheet2.")
            for item_add in self.total_item_list.keys():
                self.sheet.cell(row=start_row, column=1).value = self.total_item_list[item_add]  #来源目录
                self.sheet.cell(row=start_row, column=2).value = item_add  #商品链接
                start_row += 1
        self.save_excel(excel_name)

    def web_total_analyse(self, add='', only_category=False, webtype='m2'):
        '''pop格式:https://www.aosom.at/haus-wohnen-c130.html?page=2'''
        print("Start time:"+time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
        add=add.strip('/')
        self.category_list=self.get_category_list(add)
        for category in self.category_list:
            # print("now category:%s"%category)
            url_status=self.get_status(category)
            if only_category:
                continue
            if url_status != 200:
                continue
            category_item_total_count = self.get_total_count(category, webtype)
            if category_item_total_count:
                print("This category totally get %s item count." %category_item_total_count)
            category_item_total_count=int(category_item_total_count)
            if webtype=='pop':
                if 0 < category_item_total_count <= 21:
                    self.item_list = self.get_item_list(category)
                    for item in self.item_list:
                        item = add + item
                        if item not in self.total_item_list:
                            self.total_item_list.append(item)
                            # print("item address:%s"%item)
                            self.get_picture_address(item, True)
                elif category_item_total_count > 21:
                    get_time = category_item_total_count // 21 + 1
                    for i in range(get_time):
                        self.item_list = self.get_item_list(category + "?page=%s" % (i + 1))
                        for item in self.item_list:
                            item = add + item
                            if item not in self.total_item_list:
                                self.total_item_list.append(item)
                                # print("item address:%s"%item)
                                self.get_picture_address(item, True)
                continue
            if 0 < category_item_total_count <= 100:
                self.item_list=self.get_item_list(category+"?product_list_limit=100")
                for item in self.item_list:
                    item=add+item
                    if item not in self.total_item_list:
                        self.total_item_list.append(item)
                    # print("item address:%s"%item)
                        self.get_picture_address(item, True)
            elif category_item_total_count > 100:
                get_time=category_item_total_count//100+1
                for i in range(get_time):
                    self.item_list = self.get_item_list(category + "?product_list_limit=100&p=%s"%(i+1))
                    for item in self.item_list:
                        item = add + item
                        if item not in self.total_item_list:
                            self.total_item_list.append(item)
                            # print("item address:%s"%item)
                            self.get_picture_address(item, True)
                    # if jpglink:
                    #     for jpg_add in jpglink:
                    #         if jpg_add not in self.total_jpg_list:
                    #             self.total_jpg_list.append(jpg_add)
        print("Finish time:"+time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))

    def test_page_category(self,add=''):
        '''多线程获取'''
        print("Start time:" + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
        add = add.strip('/')
        # self.category_list = self.get_category_list(add)
        # for category in self.category_list:
        #     threads= [threading.Thread(target=self.get_status, args=(category,)) for _ in range(5)]
        thread = threading.Thread(target=self.get_category_list, args=(add,))  # 线程负责抓取列表url
        thread.start()
        time.sleep(3)
        html_thread = []
        for i in range(1, 6):
            thread2 = threading.Thread(target=self.get_status, args=(i,))
            html_thread.append(thread2)  # 开5个线程抓取页面status
        for thread_going in html_thread:
            thread_going.start()  # 等待所有线程结束，thread.join()函数代表子线程完成之前，其父进程一直处于阻塞状态。
        thread.join()
        print("The put queue thread is finished.")
        for thread_going in html_thread:
            thread_going.join()
            print("%s finished" % thread_going)
        print("The category analyse thread are finished.")
        print("Finish time:" + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))

    def get_page_canonical(self, add=''):
        '''获取页面对应的canonical的地址'''
        page_result =self.sessions.get(add)
        match = "(?<=\"canonical\" href=\")[a-zA-Z0-9\/~\-\.\:]*(?=\")"
        canonical_url=re.findall(match, page_result.text)
        if canonical_url:
        # print(canonical_url[0])
            return canonical_url[0]
        return False

    def match_canonical(self, add=''):
        '''匹配目标add和其得到的canonical是否相同'''
        add=add.strip('/')
        page_canonical=self.get_page_canonical(add)
        if page_canonical:
            page_canonical= page_canonical.strip('/')
            if page_canonical != add:
                if self.total_item_list and add in self.total_item_list.keys:
                    self.sheet.cell(row=self.start_row, column=1).value =self.total_item_list[add]
                else:
                    self.sheet.cell(row=self.start_row, column=1).value = self.father_url
                self.sheet.cell(row=self.start_row, column=2).value = add
                self.sheet.cell(row=self.start_row, column=3).value = 200
                self.sheet.cell(row=self.start_row, column=4).value = page_canonical
                self.start_row+=1
                print("add:%s canonical url:%s" %(add, page_canonical))

    def set_item_queue(self):
        '''将商品地址写入到队列中'''
        for item_add in self.total_item_list.keys():
            self.item_queue.put(item_add)

    def item_canonical_analyse(self):
        '''线程分析队列中的商详页'''
        while not self.item_queue.empty():
            add = self.item_queue.get()
            self.match_canonical(add)

    def test_page_canonical(self,add=''):
        '''多线程分析主页的链接，得到类目列表'''
        print("Start time:" + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
        excel_name=self.create_excel()
        self.read_excel(excel_name, 1)
        add = add.strip('/')
        # self.category_list = self.get_category_list(add)
        # for category in self.category_list:
        #     threads= [threading.Thread(target=self.get_status, args=(category,)) for _ in range(5)]
        thread = threading.Thread(target=self.get_category_list, args=(add, True,))  # 线程负责抓取列表url
        thread.start()
        time.sleep(3)
        html_thread = []
        self.father_url = add
        for i in range(1, 6):
            thread2 = threading.Thread(target=self.get_status_total, args=(False,))
            html_thread.append(thread2)  # 开5个线程抓取页面status
        for thread_going in html_thread:
            thread_going.start()  # 等待所有线程结束，thread.join()函数代表子线程完成之前，其父进程一直处于阻塞状态。
        thread.join()
        print("The put queue thread is finished.")
        for thread_going in html_thread:
            thread_going.join()
            print("%s finished." % thread_going)
        print("The category analyse thread are finished.")   #完成后得到了类目列表
        # print(self.category_list)
        self.save_excel(excel_name)   #第一阶段，主页分析，将结果写入excel后保存
        self.get_total_item_list(add, 'm2')
        self.write_item_list(excel_name)  #第二阶段，将商品地址和其对应的来源类目地址写入excel并保存
        self.read_excel(excel_name, 1)
        thread3= threading.Thread(target=self.set_item_queue)
        thread3.start()
        time.sleep(3)
        item_analyse_thread = []
        for i in range(1, 10):
            thread4 = threading.Thread(target=self.item_canonical_analyse)
            item_analyse_thread.append(thread4)  # 开10个线程抓取item canonical
        for thread_going in item_analyse_thread:
            thread_going.start()  # 等待所有线程结束，thread.join()函数代表子线程完成之前，其父进程一直处于阻塞状态。
        thread3.join()
        print("The put item queue thread is finished.")
        for thread_going in item_analyse_thread:
            thread_going.join()
            print("%s finished." % thread_going)
        print("The item canonical analyse thread are finished.")   #完成后得到了类目列表
        self.save_excel(excel_name)  #第三阶段，把商品地址逐个分析canonical,并写入excel
        print("Finish time:" + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))


class analyse(threading.Thread):
    def __init__(self, queue):
        threading.Thread.__init__(self)
        self.queue = queue
        self.thread_stop = False

    def run(self):
        while not self.thread_stop:
            print("thread%d %s: waiting for test" % (self.ident, self.name))

    def test_page_category(self,add=''):
        print("Start time:" + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
        add = add.strip('/')
        # self.category_list = self.get_category_list(add)
        # for category in self.category_list:
        for i in range(2):
            t = threading.Thread(target=self.get_category_list, args=(add,))
            t.start()
        for i in range(5):
            v=threading.Thread(target=self.get_status, args=(i,))
            v.start()
        print("Finish time:" + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))



if __name__ == '__main__':
    test = do_test()
    sys.stdout = output_print
    # test.get_picture_address('https://www.aosom.ca/plap/outsunny-rattan-sofa-cushion-cover-replacement-only-polyester-red~4661.html', False)
    # test.get_category_list('https://www.aosom.at/')
    # test.get_total_count('https://www.aosom.at/haus-wohnen-c130.html','pop')
    # test.web_total_analyse('https://www.aosom.it/', True, 'pop')
    # test.test_page_category('https://www.aosom.it/')
    test.test_page_canonical("https://www.aosom.ca/")

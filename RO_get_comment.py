# -*- coding: utf-8 -*-
from selenium import webdriver
from requests_html import HTMLSession
from selenium.webdriver import ActionChains  #处理鼠标悬停事件
from selenium.webdriver import TouchActions
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.common.by import By   #By类：定位元素
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.keys import Keys
import sys
import re
import os
import time,datetime
import openpyxl
from openpyxl.drawing.image import Image
from root_scripts.common import Common
from PIL import Image as pilimage
from requests_toolbelt import MultipartEncoder  #上传文件模块
from root_scripts.common import output_print
#Python3字符串默认编码unicode
libpath = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if not libpath in sys.path:
    sys.path.append(libpath)

class google_search_test(Common):
    # output_print=output_print()
    def __init__(self, logname):
        Common.__init__(self, logname)
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.212 Safari/537.36",
            "content-type": "application/x-www-form-urlencoded; charset=UTF-8"}  # "User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Safari/537.36",
        self.sessions = HTMLSession()
        self.sessions.headers = headers

    def select_excel_file(self):
        # root=tk.Tk()
        # root.withdraw()
        # filepath=filedialog.askopenfilename()  #获取文件名
        filepath = self.select_file()
        # print(filepath)
        return filepath

    def read_excel(self, excel_name='', sheetnum=1):
        '''预读取excel，方便后面写入'''
        self.wb = openpyxl.load_workbook(excel_name)
        sheetname = self.wb.sheetnames
        print("Open excel %s sheet:%s" %(excel_name, sheetnum))
        self.sheet = self.wb.get_sheet_by_name(sheetname[sheetnum-1])

    def get_url_list(self):
        rows=self.sheet.max_row
        # keyword_list=[]
        # columns = self.sheet.max_column
        url_list=[]
        for i in range(rows):
            item={'sku':'', 'add':''}
            if self.sheet.cell(row=1+i, column=3).value and 'http' in self.sheet.cell(row=1+i, column=3).value:   #只有存在关键词的才记录
                # keyword_list.append(self.sheet.cell(row=1+i, column=2).value)
                # print(keyword_list)
                item['sku']=self.sheet.cell(row=1+i, column=1).value
                item['add'] = self.sheet.cell(row=1 + i, column=3).value
                url_list.append(item)
        self.wb.close()
        return url_list
                # self.queue.put(self.sheet.cell(row=1+i, column=2).value)

    def get_page(self, add='https://www.emag.ro/motocicleta-electrica-pentru-copii-homcom-rosu-negru-102-x-53-x-66-cm-5-ani-301-043rd/pd/D3KCXZBBM/?X-Search-Id=2e23cce52f702ac35343&X-Product-Id=42205657&X-Search-Page=1&X-Search-Position=2&X-Section=search&X-MB=0&X-Search-Action=view'):
        page_result = self.sessions.get(add)
        writer=page_result.html.xpath('//p[@class="product-review-author mrg-top-xs semibold"]/text()')
        write_time = page_result.html.xpath('//p[@class="small text-muted mrg-sep-none"]/text()')
        star = page_result.html.xpath('//div[@class="star-rating-container mrg-btm-xs"]/div/div')
        comment_lens=len(star)
        for i in range(comment_lens):
            print(star[i].html)
            if '100%' in star[i].html:
                star[i]='5 stars'
            elif '80%' in star[i].html:
                star[i]='4 stars'
            elif '60%' in star[i].html:
                star[i]='3 stars'
            elif '40%' in star[i].html:
                star[i]='2 stars'
            elif '20%' in star[i].html:
                star[i]='1 stars'
        comment_title=page_result.html.xpath('//h3[@class="product-review-title"]/a/text()')
        comment_body= page_result.html.xpath('//div[@class="mrg-btm-xs js-review-body review-body-container"]/text()')
        return(writer,write_time,star,comment_title,comment_body)

    def create_excel(self):
        '''根据执行时间新建excel,表名中带国家名字'''
        excel_new=openpyxl.Workbook()
        sheet1=excel_new.active
        sheet1.title="google_keyword_result"
        sheet1.cell(row=1, column=1).value = 'SKU'
        sheet1.cell(row=1, column=2).value = '姓名'
        sheet1.cell(row=1, column=3).value = '日期'
        sheet1.cell(row=1, column=5).value = '评论标题'
        sheet1.cell(row=1, column=4).value = '星级'
        sheet1.cell(row=1, column=6).value = '评论内容'

        t = time.strftime("%Y_%m_%d", time.localtime())  #_%H_%M
        filename = 'results/' + t +'_RO_comment.xlsx'
        excel_new.save(filename)
        return filename

    def open_excel(self, filename=''):
        print('Open excel file:%s' %filename)
        self.wb_new = openpyxl.load_workbook(filename)
        sheetname_list = self.wb_new.sheetnames
        sheetname = sheetname_list[0]
        self.sheet_new = self.wb_new.get_sheet_by_name(sheetname)

    def write_excel(self, row_number=1, sku='', writer='', write_time='', star='', comment_title='', comment_body=''):
        self.sheet_new.cell(row=row_number, column=1).value = sku
        self.sheet_new.cell(row=row_number, column=2).value= writer
        self.sheet_new.cell(row=row_number, column=3).value = write_time
        self.sheet_new.cell(row=row_number, column=4).value = star
        self.sheet_new.cell(row=row_number, column=5).value = comment_title
        self.sheet_new.cell(row=row_number, column=6).value = comment_body

    def save_excel(self,excel_name=''):
        self.wb_new.save(excel_name)

    def total_test(self):
        filepath = self.select_excel_file()
        self.read_excel(filepath)
        url_list=self.get_url_list()
        # print(url_list)
        filename = self.create_excel()
        self.open_excel(filename)
        row_number=2
        for url in url_list:
            print(url)
            sku=url['sku']
            add=url['add']
            writer, write_time, star, comment_title, comment_body=self.get_page(add)
            comment_len=len(star)
            for i in range(comment_len):
                self.write_excel(row_number,sku,writer[i],write_time[i],star[i],comment_title[i],comment_body[i])
                row_number+=1
        self.save_excel(filename)

if __name__ == '__main__':
    test = google_search_test('test')
    test.total_test()

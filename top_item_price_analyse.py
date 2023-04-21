# -*- coding: utf-8 -*-
from selenium import webdriver
from requests_html import HTMLSession
from selenium.webdriver import ActionChains  #处理鼠标悬停事件
from selenium.webdriver import TouchActions
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.common.by import By   #By类：定位元素
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.keys import Keys
import sys
import re
import os
import time,datetime
import openpyxl
from openpyxl.drawing.image import Image
from root_scripts.common import Common
from PIL import Image as pilimage
from requests_toolbelt import MultipartEncoder  #上传文件模块
from root_scripts.common import output_print
#Python3字符串默认编码unicode
libpath = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if not libpath in sys.path:
    sys.path.append(libpath)

class google_search_test(Common):
    # output_print=output_print()
    def __init__(self, logname):
        Common.__init__(self, logname)
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.212 Safari/537.36",
            "content-type": "application/x-www-form-urlencoded; charset=UTF-8"}  # "User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Safari/537.36",
        self.sessions = HTMLSession()
        self.sessions.headers = headers
        self.natual_items = []  # 用来存自然流量第一页所有商品的信息
        self.price_list = []  #用来存储当页的价格数据
        self.first_item = {'name': '', 'price': 100000.00, 'photo_add': ''} #存储当页最便宜的商品
        self.second_item = {'name': '', 'price': 100000.00, 'photo_add': ''}  # 存储当页次便宜的商品
        self.third_item = {'name': '', 'price': 100000.00, 'photo_add': ''}  # 存储当页季便宜的商品

    def set_up(self, web='https://www.google.co.uk/', device_type='html'):
        '''常见的操作元素方法如下：
        clear 清除元素的内容
        send_keys 模拟按键输入
        click 点击元素
        submit 提交表单'''
        device_type=device_type.upper()
        if device_type=='H5':
            mobile_emulation = {'deviceName': 'iPhone X'}
            options = webdriver.ChromeOptions()
            options.add_experimental_option('mobileEmulation', mobile_emulation)
        else:
            options = webdriver.ChromeOptions()
            options.add_experimental_option('debuggerAddress', '127.0.0.1:9222')  #chrome.exe --remote-debugging-port=9222 --user-data-dir="C:\selenum\AutomationProfile"
            options.add_argument("start-maximized")
        if os.path.exists("D:/python/chromedriver.exe"):
            self.browser = webdriver.Chrome("D:/python/chromedriver.exe", chrome_options=options)
        else:
            self.browser = webdriver.Chrome("C:/Users/aosom/AppData/Local/Google/Chrome/Application/chromedriver.exe",
                                            chrome_options=options)
        try:
            self.browser.implicitly_wait(10)  # 隐式等待10秒，如果在10秒内网页加载完成，则执行下一步，否则一直等到时间截止，然后执行下一步
            self.browser.get(web)   #在当前窗口加载 url
            self.logger.info("Start web:%s" %web)
            # self.browser.maximize_window()
            print("Set window to max size.")
        except TimeoutException:
            print("Can not open the web!")

    def close(self):
        self.browser.close()  #关闭当前窗口, 如果当前窗口是最后一个窗口, 浏览器将关闭
        self.browser.quit()  #关闭所有窗口并停止 ChromeDriver 的执行
        print("Close the browser.")

    def skip_search_info_page(self):
        try:
            buttons=self.browser.find_elements_by_xpath('//div[@class="VfPpkd-dgl2Hf-ppHlrf-sM5MNb"]/button')
            if buttons:
                buttons[1].click()  #点第二个我同意跳过
                time.sleep(3)
        except:
            return True

    def search(self, keyword=''):
        self.browser.find_element_by_class_name('gLFyf').clear()
        self.browser.find_element_by_class_name('gLFyf').send_keys(keyword)
        self.browser.find_element_by_class_name('gLFyf').send_keys(Keys.ENTER)
        self.logger.info("Search keyword:%s" % keyword)
        time.sleep(10)  #搜索完多等待一会

    def get_redirects_add(self, add=''):
        '''获得302跳转页面的最终页面链接'''
        # if add.split(':')[0] == 'https':
        #     result = self.sessions.get(url=add, proxies='https')
        # else:
        #     result = self.sessions.get(url=add, proxies='http')
        result = self.sessions.get(url=add)
        his_add=result.history
        # print (his_add)
        return his_add[-1].headers["location"]

    def deal_name(self, name=''):
        '''处理字符串，去掉中间特殊的字符和空格'''
        deal_name="".join(filter(str.isalnum, name))
        # print(deal_name)
        return deal_name

    def test(self):
        price='US$2,499.00'
        digit_price_list = re.findall('\d?,?\d+\.?,?\d+',
                                      price)  # 最复杂的价格格式€1,938.81 '1,239.99' £98.99 £8.99 US$148.74 aZK3gc 21,99 €  8239,00 €\d?,?\d+(\.|,)\d+
        if len(digit_price_list) > 1:
            digit_price = digit_price_list[0]
        else:
            digit_price = digit_price_list
        print(digit_price)

    def get_page_natual_price(self):
        '''获取页面natual部分关键信息,根据关键字的个数来控制抓的条数'''
        self.natual_items=[]  #每一轮执行前都清除上一轮的数组
        self.price_list=[]
        self.first_item['price'] = 100000.00
        self.second_item['price'] = 100000.00
        self.third_item['price'] = 100000.00
        aditems = self.browser.find_elements_by_class_name('sh-dgr__grid-result')
        total_count = len(aditems)
        if total_count > 40:
            loop_count = 40
        else:
            loop_count = total_count
        print("Totally find %s matched natual items,Get %s items price." % (total_count, loop_count))
        price_three=[100000.00, 100000.00, 100000.00]  #用来保存最小的3个单价
        for i in range(loop_count):
            temp_item={'name': '', 'price': 100000.00, 'photo_add': ''}  #,'source': '', 'address': ''
            if self.iselementexits_by_classname(self.browser, 'A2sOrd'):
                name = aditems[i].find_element_by_class_name('A2sOrd').text
            elif self.iselementexits_by_classname(self.browser, 'Xjkr3b'):
                name = aditems[i].find_element_by_class_name('Xjkr3b').text
            else:
                name = ''
            if self.iselementexits_by_classname(self.browser, 'a8Pemb'):
                price=aditems[i].find_element_by_class_name('a8Pemb').text
                print("Get price:%s" %price)
                if price:
                    digit_price=re.findall('\d?,?\d+\.?,?\d+', price)[0]   #最复杂的价格格式€1,938.81 '1,239.99' £98.99 £8.99 US$148.74 aZK3gc 21,99 €  8239,00 €\d?,?\d+(\.|,)\d+
                    # if len(digit_price_list)>1:
                    #     digit_price=digit_price_list[0]
                    # else:
                    #     digit_price=digit_price_list
                    # if len(digit_price)>1:
                    #     digit_price=digit_price[0]
                    # print(digit_price)
                    if ',' in digit_price:
                        # print('remove the , in price string:%s.' %digit_price)
                        digit_price = digit_price.replace(',', '')
                    if '.' in digit_price:
                        # print('remove the . in price string:%s.' %digit_price)
                        digit_price = digit_price.replace('.', '')
                    # print(digit_price)
                    # for i in price:
                    #     if i.isdight:
                    #         digit_price+=i
                    digit_price=float(digit_price)*0.01
                    self.price_list.append(digit_price)
                else:
                    price = ''
                    digit_price = 100000.00
            else:
                price = ''
                digit_price = 100000.00
            if digit_price != 100000.00:
                for count in range(3):
                    if digit_price < price_three[count]:
                        price_three[count]=digit_price
                        google_add=aditems[i].find_element_by_class_name('xCpuod').get_attribute('href')  #'https://www.google.com'+
                        try:
                            aditems[i].find_element_by_class_name("ArOc1c").click()  #先点击展开商品的扩展框
                            time.sleep(4)
                        except:
                            pass
                        if "https://www.google.com" in google_add:
                            google_add_part=google_add.replace('https://www.google.com', '')
                        else:
                            google_add_part=google_add
                        photo_xpath='//a[@href="' + google_add_part +'"]/div[1]/div[1]'
                        if self.iselementexits_by_xpath(self.browser, photo_xpath):
                            photo_add_path = aditems[i].find_element_by_xpath(photo_xpath).get_attribute('innerHTML')
                            # print(photo_add_path)
                            match = "(?<=src=\")https.*?(?=\")"
                            photo_add = re.findall(match, photo_add_path)
                        else:
                            photo_add=''
                        if len(photo_add)==0:
                            photo_add=''
                        else:
                            photo_add=photo_add[0]
                        print('loop %s photo add is :%s' %(i+1,photo_add))
                        break
            # save_name = self.deal_name(name)
            # self.download_image(photo_add, save_name)
            # try:
            #     add=self.get_redirects_add(google_add)
            # except:
            #     add=google_add
            temp_item['name'] = name
            temp_item['price'] = digit_price
            # temp_item['source'] = source
            # temp_item['address'] = add
            temp_item['photo_add'] = photo_add
            self.natual_items.append(temp_item)
        for item in self.natual_items:
            if item['price'] <= self.first_item['price']:   #价格比最小还要小，把价格排序往后赋值一遍
                self.third_item['price'] = self.second_item['price']
                self.third_item['name'] = self.second_item['name']
                self.third_item['photo_add'] = self.second_item['photo_add']
                self.second_item['price'] = self.first_item['price']
                self.second_item['name'] = self.first_item['name']
                self.second_item['photo_add'] = self.first_item['photo_add']
                self.first_item['price'] = item['price']
                self.first_item['name'] = item['name']
                self.first_item['photo_add'] = item['photo_add']
            elif item['price'] <= self.second_item['price']: #价格比次小还要小，把价格排序往后赋值一遍
                self.third_item['price'] = self.second_item['price']
                self.third_item['name'] = self.second_item['name']
                self.third_item['photo_add'] = self.second_item['photo_add']
                self.second_item['price'] = item['price']
                self.second_item['name'] = item['name']
                self.second_item['photo_add'] = item['photo_add']
            elif item['price'] < self.third_item['price']: #价格比季小还要小
                self.third_item['price'] = item['price']
                self.third_item['name'] = item['name']
                self.third_item['photo_add'] = item['photo_add']
        save_name = self.deal_name(self.third_item['name'])
        self.download_image(self.third_item['photo_add'], save_name)
        save_name = self.deal_name(self.second_item['name'])
        self.download_image(self.second_item['photo_add'], save_name)
        save_name = self.deal_name(self.first_item['name'])
        self.download_image(self.first_item['photo_add'], save_name)

    def get_price_rank(self, price=0.00):
        '''得到价格在所有价格列表中的低价排名'''
        rank = 1
        for p in self.price_list:
            if p < price:
                rank += 1
        return rank

    def get_avg_price(self):
        if self.price_list:
            average=sum(self.price_list)/len(self.price_list)
            return round(average, 2)
        return 0

    def create_excel(self, country='CN'):
        '''根据执行时间新建excel,表名中带国家名字'''
        excel_new=openpyxl.Workbook()
        sheet1=excel_new.active
        sheet1.title="google_keyword_result"
        sheet1.cell(row=1, column=1).value = 'SKU'
        sheet1.cell(row=1, column=2).value = '售价'
        sheet1.cell(row=1, column=3).value = '品名'
        sheet1.cell(row=1, column=4).value = '关键词'
        sheet1.cell(row=1, column=5).value = '价格排名'
        sheet1.cell(row=1, column=6).value = '搜索页平均价格'
        sheet1.cell(row=1, column=7).value = '最便宜商品名'
        sheet1.cell(row=1, column=8).value = '最便宜商品图片'
        sheet1.cell(row=1, column=9).value = '最便宜商品价格'
        sheet1.cell(row=1, column=10).value = '次便宜商品名'
        sheet1.cell(row=1, column=11).value = '次便宜商品图片'
        sheet1.cell(row=1, column=12).value = '次便宜商品价格'
        sheet1.cell(row=1, column=13).value = '季便宜商品名'
        sheet1.cell(row=1, column=14).value = '季便宜商品图片'
        sheet1.cell(row=1, column=15).value = '季便宜商品价格'
        t = time.strftime("%Y_%m_%d", time.localtime())  #_%H_%M
        filename = 'results/' + t + '_'+country+'_top_items_price_rank.xlsx'
        excel_new.save(filename)
        return filename

    def write_excel(self, filename='', sku='', price=0.00, name='', keyword='', rank=1, average=0.00):
        '''将数组中存放的内容写入excel,需求纵向排序在下方追加写入'''
        wb = openpyxl.load_workbook(filename)
        sheetname_list = wb.sheetnames
        sheetname = sheetname_list[0]
        sheet = wb.get_sheet_by_name(sheetname)
        # columns = sheet.max_column  #横向扩展，获取最大的列数
        # if columns==1:
        #     columns=0
        columns = 0
        rows=sheet.max_row
        # if rows==1:
        #     rows=1
        sheet.cell(row=rows+1, column=1).value = sku
        sheet.cell(row=rows+1, column=2).value = price
        sheet.cell(row=rows+1, column=3).value = name
        sheet.cell(row=rows+1, column=4).value = keyword
        sheet.cell(row=rows + 1, column=5).value = rank
        sheet.cell(row=rows + 1, column=6).value = average
        if self.first_item['price']!=100000.00:
            sheet.cell(row=rows+1, column=7).value =self.first_item['name']
            sheet.cell(row=rows+1, column=9).value =self.first_item['price']
            photo_name=self.deal_name(self.first_item['name'])
            image_path='results/' + photo_name + '.jpg'
            if os.path.exists(image_path):
                image = pilimage.open(image_path)
                image.save(image_path, 'JPEG')
                img = Image(image_path)  # 选择你的图片
                img.width, img.height = 75, 75
                sheet.add_image(img, 'H' + str(rows+1))
                sheet.row_dimensions[rows+1].height = 70  # 行高
            else:
                sheet.cell(row=rows+1, column=8).value = self.first_item['photo_add']
        if self.second_item['price']!=100000.00:
            sheet.cell(row=rows+1, column=10).value =self.second_item['name']
            sheet.cell(row=rows+1, column=12).value =self.second_item['price']
            photo_name=self.deal_name(self.second_item['name'])
            image_path='results/' + photo_name + '.jpg'
            if os.path.exists(image_path):
                image = pilimage.open(image_path)
                image.save(image_path, 'JPEG')
                img = Image(image_path)  # 选择你的图片
                img.width, img.height = 75, 75
                sheet.add_image(img, 'K' + str(rows+1))
                sheet.row_dimensions[rows+1].height = 70  # 行高
            else:
                sheet.cell(row=rows+1, column=11).value = self.second_item['photo_add']
        if self.third_item['price']!=100000.00:
            sheet.cell(row=rows+1, column=13).value =self.third_item['name']
            sheet.cell(row=rows+1, column=15).value =self.third_item['price']
            photo_name=self.deal_name(self.third_item['name'])
            image_path='results/' + photo_name + '.jpg'
            if os.path.exists(image_path):
                image = pilimage.open(image_path)
                image.save(image_path, 'JPEG')
                img = Image(image_path)  # 选择你的图片
                img.width, img.height = 75, 75
                sheet.add_image(img, 'N' + str(rows+1))
                sheet.row_dimensions[rows+1].height = 70  # 行高
            else:
                sheet.cell(row=rows+1, column=14).value = self.third_item['photo_add']
        wb.save(filename)

    def download_image(self, add='', name=''):
        '''下载图片'''
        try:
            res = self.sessions.get(add)
            file_name = 'results/'+name+'.jpg'
            with open(file_name, 'wb') as f:
                f.write(res.content)
                print("Download photo %s.jpg succeed!" % name)
            f.close()
        except:
            print("Download photo %s.jpg failed!"%name)
            return False

    def get_country_list(self):
        '''读取excel表中国家列表'''
        wb = openpyxl.load_workbook('TOP产品价格排名.xlsx')
        sheetname_list = wb.sheetnames
        # print(sheetname_list)
        return sheetname_list

    def select_country(self, country_list=[]):
        '''从国家列表选择国家'''
        '''将国家列表展示出来，并选择'''
        for i in range(len(country_list)):
            print("%s:%s"%(i+1,country_list[i]))
        country_num=input("PLS select the country ID:")
        country_num=int(country_num)
        if 0< country_num <=len(country_list):
            print("You have select:%s"%country_list[country_num-1])
            return country_list[country_num-1]
        else:
            print("Country ID error!")
            self.select_country(country_list)

    def get_keyword_list(self, country_name=''):
        wb = openpyxl.load_workbook('TOP产品价格排名.xlsx')
        sheet = wb.get_sheet_by_name(country_name)
        rows=sheet.max_row
        keyword_list=[]
        columns = sheet.max_column
        for i in range(columns):
            value = sheet.cell(row=1, column=i + 1).value
            if value == 'SKU':
                sku_column = i + 1
            elif value == '售价'or value == 'price' or value == '价格':
                price_column = i + 1
            elif value == '品名':
                name_column = i + 1
            elif value == '关键词' or value == 'keyword':
                keyword_column = i + 1
        for i in range(rows):
            item_info = {'sku': '', 'price': 0.00, 'name': '', 'keyword': ''}
            if sheet.cell(row=2 + i, column=keyword_column).value:   #只有存在关键词的才记录
                item_info['sku']=sheet.cell(row=2 + i, column=sku_column).value   #从第二行开始
                item_info['price'] = sheet.cell(row=2 + i, column=price_column).value if sheet.cell(row=2 + i, column=price_column).value else 0.00
                item_info['name'] = sheet.cell(row=2 + i, column=name_column).value if sheet.cell(row=2 + i, column=name_column).value else ''
                item_info['keyword'] = sheet.cell(row=2 + i, column=keyword_column).value if sheet.cell(row=2 + i, column=keyword_column).value else ''
                keyword_list.append(item_info)
                # print(keyword_list)
        return keyword_list

    def upload_file(self, filename='', filepath=''):
        '''上传文件'''
        headers={
                 'Content-Type': 'multipart/form-data; boundary=----WebKitFormBoundaryHF2C97Tky0OzQDQm',
                }
        # data = {
        #     'name': filename,
        #     # 'Content-Disposition': 'form-data',
        #     # 'Content-Type': 'multipart/form-data',
        #     # 'boundary':'',
        #     # "Content-Length":'',
        #     # 'Content-Type': 'application/vnd.ms-excel'
        # }
        data =MultipartEncoder(fields={'file': (filename, open(filepath, 'rb'), 'application/vnd.ms-excel'),}  #'identityType':identityType
                               ,boundary='----WebKitFormBoundaryHF2C97Tky0OzQDQm')
        # with open(filepath, 'rb') as files:
        # files = {'file': open(filepath, 'rb')}
        # files = {'file': (filename, open(filepath, 'rb'), 'application/vnd.ms-excel', {'Expires': '0'})}
        print('Upload file %s'%filename)
        response = self.sessions.post('https://www.aosom.co.uk/aosom_ca/product/infoupload', headers=headers, data=data)#files=files
        print(response.text)
        return response

    def delete_photo(self):
        '''清空爬下来的所有图片'''
        files=os.listdir('results/')
        for file in files:
            if '.jpg' in file:
                os.remove('results/' + file)
        # print(files)

    def total_test(self):
        country_list = self.get_country_list()
        country = self.select_country(country_list)
        keyword_list = self.get_keyword_list(country)
        # while True:
            # if self.check_time('22'):
        filename=self.create_excel(country)
        # self.set_up('https://www.google.com/search?q=hollywoodschaukeln&tbm=shop&ei=HnPyYOfTIeGDlQfHwZSQBA&oq=hollywoodschaukeln&gs_lcp=Cgtwcm9kdWN0cy1jYxADULEqWNgqYOsqaABwAHgAgAEAiAEAkgEAmAEAoAEBwAEB&sclient=products-cc&ved=0ahUKEwjn2tPot-nxAhXhQeUKHccgBUIQ4dUDCAs&uact=5')
        self.set_up('https://www.google.com/search?q=aosom&source=lnms&tbm=shop')
        self.skip_search_info_page()
        keyword_count=len(keyword_list)
        keyword_have_set=[]  #将已统计过的关键词放入
        for i in range(keyword_count):   #每一次执行遍历一遍所有的关键词列，优先把相同关键词的产品数据填入
            if keyword_list[i]['keyword'] not in keyword_have_set:
                keyword_have_set.append(keyword_list[i]['keyword'])
                self.search(keyword_list[i]['keyword'])
                self.get_page_natual_price()
                rank= self.get_price_rank(keyword_list[i]['price'])
                average=self.get_avg_price()
                self.write_excel(filename, keyword_list[i]['sku'], keyword_list[i]['price'], keyword_list[i]['name'], keyword_list[i]['keyword'], rank, average)
                time.sleep(1)
                for j in range(i+1,keyword_count): #把当前关键词下面的都遍历一遍
                    if keyword_list[j]['keyword'] == keyword_list[i]['keyword']:
                        rank = self.get_price_rank(keyword_list[j]['price'])
                        self.write_excel(filename, keyword_list[j]['sku'], keyword_list[j]['price'], keyword_list[j]['name'], keyword_list[j]['keyword'], rank, average)
                        time.sleep(1)
        self.close()
        filename_short=filename.replace('results/', '')
        self.upload_file(filename_short, filename)
        self.delete_photo()



if __name__ == '__main__':
    test = google_search_test('test')
    test.total_test()

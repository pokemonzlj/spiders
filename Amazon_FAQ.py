# -*- coding: utf-8 -*-
from selenium import webdriver
from selenium.webdriver import ActionChains  #处理鼠标悬停事件
from selenium.webdriver import  TouchActions
from selenium.common.exceptions import TimeoutException, NoSuchElementException, ElementClickInterceptedException, ElementNotInteractableException
from selenium.webdriver.common.by import By   #By类：定位元素
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.keys import Keys
import sys
import os
import time,datetime
import random
import openpyxl
from rootscripts.common import Common

libpath = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if not libpath in sys.path:
    sys.path.append(libpath)

class FAQ(Common):
    def __init__(self,  log_name):
        Common.__init__(self, log_name)
        self.current_site=''

    def set_up(self, web='https://www.amazon.co.uk/', device_type='html'):
        '''常见的操作元素方法如下：
        clear 清除元素的内容
        send_keys 模拟按键输入
        click 点击元素
        submit 提交表单'''
        device_type=device_type.upper()
        if device_type=='H5':
            mobile_emulation = {'deviceName': 'iPhone X'}
            options = webdriver.ChromeOptions()
            options.add_experimental_option('mobileEmulation', mobile_emulation)
        else:
            options = webdriver.ChromeOptions()
            options.add_experimental_option('debuggerAddress', '127.0.0.1:9222')  #chrome.exe --remote-debugging-port=9222 --user-data-dir="C:\selenum\AutomationProfile"
            options.add_argument("start-maximized")
        self.browser=webdriver.Chrome("C:/Users/aosom/AppData/Local/Google/Chrome/Application/chromedriver.exe", chrome_options=options)#firefox_profile=profile,
        try:
            self.browser.implicitly_wait(10)  # 隐式等待10秒，如果在10秒内网页加载完成，则执行下一步，否则一直等到时间截止，然后执行下一步
            self.browser.get(web)   #在当前窗口加载 url
            self.logger.info("Start web:%s" %web)
            # self.browser.maximize_window()
            print("Set window to max size.")
        except TimeoutException:
            print("Can not open the web!")

    def close(self):
        self.browser.close()  #关闭当前窗口, 如果当前窗口是最后一个窗口, 浏览器将关闭
        self.browser.quit()  #关闭所有窗口并停止 ChromeDriver 的执行
        print("Close the browser.")

    def get_country_list(self):
        '''读取excel表中国家列表'''
        wb = openpyxl.load_workbook('FAQ亚马逊.xlsx')
        sheetname_list = wb.sheetnames
        # print(sheetname_list)
        return sheetname_list

    def select_country(self, country_list=[]):
        '''从国家列表选择国家'''
        '''将国家列表展示出来，并选择'''
        for i in range(len(country_list)):
            print("%s:%s"%(i+1,country_list[i]))
        country_num=input("PLS select the country ID:")
        country_num=int(country_num)
        if 0< country_num <=len(country_list):
            print("You have select:%s"%country_list[country_num-1])
            return country_list[country_num-1]
        else:
            print("Country ID error!")
            self.select_country(country_list)

    def get_asin_list(self, country_name=''):
        wb = openpyxl.load_workbook('FAQ亚马逊.xlsx')
        sheet = wb.get_sheet_by_name(country_name)
        row=sheet.max_row
        asin_list=[]
        for i in range(row):
            if sheet.cell(row=2+i, column=2).value:  #从第二行开始
                asin_list.append(sheet.cell(row=2+i, column=2).value)
        # print(keyword_list)
        return asin_list

    def get_site_url(self, site='uk'):
        site=site.lower()
        if site=='us':
            self.current_site='https://www.amazon.com/'
        elif site=='uk':
            self.current_site='https://www.amazon.co.uk/'
        else:
            self.current_site='https://www.amazon.'+site+'/'
        return self.current_site

    def get_site_postcode(self, site='uk'):
        site = site.lower()
        if site == 'us':
            postcode = '10041'
        elif site == 'uk':
            postcode = 'NW1 6XE'
        elif site == 'ca':
            postcode = 'M4Y1M7'
        elif site == 'it':
            postcode = '00144'
        elif site == 'es':
            postcode = '28000'
        elif site == 'fr':
            postcode = '75020'
        elif site == 'de':
            postcode = '10115'
        else:
            return False
        return postcode

    def accept_cookie(self):
        if self.iselementexist_by_id(self.browser, 'sp-cc-accept'):
            self.browser.find_element_by_id('sp-cc-accept').click()
            print('Click accept cookies.')
            self.delay(1)

    def switch_location(self, postcode='NW1 6XE'):
        self.browser.find_element_by_id('glow-ingress-block').click()
        self.delay(5)
        if self.iselementexist_by_classname(self.browser, 'GLUX_Full_Width'):
            self.browser.find_element_by_class_name('GLUX_Full_Width').send_keys(postcode)
            buttons=self.browser.find_elements_by_class_name('a-button-input')
            buttons[-1].click()
            self.delay(3)

    def get_url(self, Asin='B08KRSQ1Y1', page=1):
        address=self.current_site + 'ask/questions/asin/' + Asin + '/'+ str(page) +'?isAnswered=true'
        print("Current address is:%s" %address)
        return address

    def get_questions_count(self):
        if self.iselementexist_by_classname(self.browser, 'askPaginationHeaderMessage'):
            count_text=self.browser.find_element_by_class_name('askPaginationHeaderMessage').text
            telist=count_text.split()
            for te in telist:
                if te.isdigit():
                    print("This commodity get %s Q&A."%te)
                    return te
        print("This commodity has no Q&A.")
        return False

    def get_address_list(self, asin=''):
        count=self.get_questions_count()
        if count:
            count = int(count)
            if count%10 == 0:
                page_count = count//10
            else:
                page_count = count//10+1
            if page_count == 1:
                return 1
            add_list=[]
            for i in range(2, page_count+1):
                add = self.get_url(asin, i)
                add_list.append(add)
            return add_list
        return False

    def get_all_question(self):
        questions=self.browser.find_elements_by_xpath("//a[@class='a-link-normal']/span[contains(@data-ask-no-op, 'top-question-text-click')]")
        questions_text=[]
        for q in questions:
            questions_text.append(q.text)
        # print(questions_text)
        return questions_text

    def get_all_answer(self):
        answers = self.browser.find_elements_by_xpath("//div[@class='a-fixed-left-grid-col a-col-right']/span[1]")
        answers_text = []
        for a in answers:
            answers_text.append(a.text)
        # print(answers_text)
        return answers_text

    def match_QandA(self):
        questions = self.get_all_question()
        answers = self.get_all_answer()
        q_lens=len(questions) if questions else 0
        a_lens=len(answers) if answers else 0
        if q_lens != a_lens:
            self.logger.warning("回答和问题个数不匹配!")
            return {}
        qanda_dict={}
        for i in range(q_lens):
            qanda_dict[questions[i]]=answers[i]
        print(qanda_dict)
        return qanda_dict

    def create_excel(self, country='CN'):
        '''根据执行时间新建excel,表名中带国家名字'''
        excel_new=openpyxl.Workbook()
        sheet1=excel_new.active
        sheet1.title="FAQ"
        sheet1.cell(row=1, column=1).value = 'Asin'
        sheet1.cell(row=1, column=2).value = 'Question'
        sheet1.cell(row=1, column=3).value = 'Answer'
        t = time.strftime("%Y_%m_%d", time.localtime())  #_%H_%M
        filename = 'results/' + t + '_'+country+'_FAQ.xlsx'
        excel_new.save(filename)
        return filename

    def write_excel(self, filename='', asin='', qanda={}):
        '''将字典中存放的内容写入excel,需求纵向排序在下方追加写入'''
        wb = openpyxl.load_workbook(filename)
        sheetname_list = wb.sheetnames
        sheetname = sheetname_list[0]
        sheet = wb.get_sheet_by_name(sheetname)
        columns = 0
        rows=sheet.max_row
        # if rows==1:
        #     rows=0
        start_row=rows+1  #从下一排开始
        for key in qanda.keys():
            sheet.cell(row=start_row, column=1).value = asin
            sheet.cell(row=start_row, column=2).value = key
            sheet.cell(row=start_row, column=3).value = qanda[key]
            start_row+=1
        wb.save(filename)

    def total_catch(self):
        country_list = self.get_country_list()
        country = self.select_country(country_list)
        asin_list = self.get_asin_list(country)
        url = self.get_site_url(country)
        filename = self.create_excel(country)
        self.set_up(url)
        self.accept_cookie()
        postcode=self.get_site_postcode(country)
        self.switch_location(postcode)
        for asin in asin_list:
            add=self.get_url(asin)
            self.browser.get(add)
            analyse_result=self.get_address_list(asin)
            if analyse_result:
                qanda_dict=self.match_QandA()
                self.write_excel(filename, asin, qanda_dict)
                self.delay(2)
                if analyse_result != 1:
                    for add in analyse_result:
                        self.browser.get(add)
                        qanda_dict = self.match_QandA()
                        self.write_excel(filename, asin, qanda_dict)
                        self.delay(2)


if __name__ == '__main__':
    test = FAQ('faq_get')
    test.total_catch()

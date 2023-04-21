# -*- coding: utf-8 -*-
from selenium import webdriver
from selenium.webdriver import ActionChains  #处理鼠标悬停事件
from selenium.webdriver import  TouchActions
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.common.by import By   #By类：定位元素
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.keys import Keys
import sys
import re
import os
import time,datetime
import openpyxl
'''如果需要设置字体、加粗和居中等，可以这样：
# 先导入这个包
from openpyxl.styles import Color, Font, Alignment
# 然后如下设置：
# 设置表头字体居中
font = Font(name=u'宋体', bold = True)
align = Alignment(horizontal='center', vertical='center')
workSheet.cellstyle('A1', font, align)'''
from root_scripts.common import Common
from root_scripts.common import output_print
#Python3字符串默认编码unicode
libpath = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if not libpath in sys.path:
    sys.path.append(libpath)

class wayfair_test(Common):
    output_print=output_print()
    def __init__(self, logname):
        Common.__init__(self, logname)
        self.weight_dict = { }  #用来存储Weights & Dimensions相关信息
        self.infomation_dict = {}  #用来存储Specifications相关信息
        self.weight_name='' #用来存储Weights & Dimensions的text的值
        self.infomation_name=''    #用来存储Specifications的text的值
        self.web_list=[]
        self.current_type=''  #存储当前的商品类型
        self.current_url=''

    def set_up(self, web='https://www.amazon.co.uk/', device_type='html'):
        '''常见的操作元素方法如下：
        clear 清除元素的内容
        send_keys 模拟按键输入
        click 点击元素
        submit 提交表单'''
        device_type=device_type.upper()
        if device_type=='H5':
            mobile_emulation = {'deviceName': 'iPhone X'}
            options = webdriver.ChromeOptions()
            options.add_experimental_option('mobileEmulation', mobile_emulation)
        else:
            options = webdriver.ChromeOptions()
            options.add_experimental_option('debuggerAddress', '127.0.0.1:9222')  #chrome.exe --remote-debugging-port=9222 --user-data-dir="C:\selenum\AutomationProfile"
            options.add_argument("start-maximized")
        self.browser=webdriver.Chrome("C:/Users/aosom/AppData/Local/Google/Chrome/Application/chromedriver.exe", chrome_options=options)#firefox_profile=profile,
        try:
            self.browser.implicitly_wait(10)  # 隐式等待10秒，如果在10秒内网页加载完成，则执行下一步，否则一直等到时间截止，然后执行下一步
            self.browser.get(web)   #在当前窗口加载 url
            self.logger.info("Start web:%s" %web)
            # self.browser.maximize_window()
            print("Set window to max size.")
        except TimeoutException:
            print("Can not open the web!")

    def close(self):
        self.browser.close()  #关闭当前窗口, 如果当前窗口是最后一个窗口, 浏览器将关闭
        self.browser.quit()  #关闭所有窗口并停止 ChromeDriver 的执行
        print("Close the browser.")

    def search(self, item_name=''):
        '''按字符串查找商品'''
        if item_name.startswith('w'):
            if self.current_url != 'https://www.wayfair.com':
                self.browser.get('https://www.wayfair.com')
                self.current_url = 'https://www.wayfair.com'
                self.browser.implicitly_wait(10)
        else:
            if self.current_url != 'https://www.wayfair.com':
                self.browser.get('https://www.wayfair.com')
                self.current_url = 'https://www.wayfair.com'
                self.browser.implicitly_wait(10)
            # if self.current_url != 'https://www.wayfair.co.uk':
            #     self.browser.get('https://www.wayfair.co.uk')
            #     self.current_url = 'https://www.wayfair.co.uk'
            #     self.browser.implicitly_wait(10)
        # self.browser.find_element_by_class_name('SearchBar-input').clear() #直接clear无法清除
        # time.sleep(1)
        # self.browser.find_element_by_class_name('SearchBar-input').send_keys(Keys.CONTROL+'a')    #全选
        # self.browser.find_element_by_class_name('SearchBar-input').send_keys(Keys.DELETE)	# 删除，清空
        # self.browser.find_element_by_class_name('SearchBar-input').send_keys(item_name)
        # print("Input search item name:%s"%item_name)
        # self.browser.find_element_by_class_name('SearchBar-button').click()
        # print("Click search.")
        # self.browser.find_element_by_class_name('pl-TextInput-input').send_keys(Keys.CONTROL + 'a')  # 全选
        # self.browser.find_element_by_class_name('pl-TextInput-input').send_keys(Keys.DELETE)  # 删除，清空
        if self.iselementexits_by_classname(self.browser, 'pl-TextInput-icon--clear'):
            self.browser.find_element_by_class_name('pl-TextInput-icon--clear').click()
            time.sleep(1)
        # self.browser.find_element_by_class_name('pl-TextInput-input').clear()
        while not self.iselementexits_by_classname(self.browser, 'pl-TextInput-input'):
            self.logger.debug("Google robot check,wait 10 seconds")
            time.sleep(10)
        self.browser.find_element_by_class_name('pl-TextInput-input').send_keys(item_name)
        print("Input search item name:%s" % item_name)
        self.browser.find_element_by_class_name('pl-TextInput-input').send_keys(Keys.ENTER)
        print("Click search.")
        time.sleep(2)
        try:
            WebDriverWait(self.browser, 30).until(EC.presence_of_element_located((By.CLASS_NAME, "Specifications-descriptionList-cell")))
            return True
        except TimeoutException:
            return False

    def get_item_info(self):
        '''获取商品对应的信息'''
        # info_list=self.browser.find_elements_by_class_name('Specifications-descriptionList-cell')
        title_list=self.browser.find_elements_by_class_name('ProductOverviewItem-title')
        # if self.iselementexits_by_xpath(self.browser, '//div[@class="ProductWeightsDimensions-descriptionListItem"]'):
        #     if not self.browser.find_element_by_xpath('//div[@class="ProductWeightsDimensions-descriptionListItem"]').text:
        #         if self.weight_name== '':
        #             self.weight_name=title_list[1].text
        #         title_list[1].click()
        #         time.sleep(1)
        #     weight_list = self.browser.find_elements_by_xpath(
        #         '//div[@class="ProductWeightsDimensions-descriptionListItem"]')
        #     weight_input_times = len(weight_list) // 2
        #     for i in range(weight_input_times):
        #         key = weight_list[0 + 2 * i].text
        #         value = weight_list[1 + 2 * i].text
        #         if not self.weight_dict.__contains__(key):
        #             self.weight_dict[key] = value
        if self.iselementexits_by_xpath(self.browser, '//div[@class="Specifications-descriptionList-cell"]'):
            if not self.browser.find_element_by_xpath('//div[@class="Specifications-descriptionList-cell"]').text:
                if self.infomation_name== '':
                    self.infomation_name=title_list[2].text
                title_list[2].click()
                # print(title_list[2].text)
                time.sleep(1)
                info_list = self.browser.find_elements_by_xpath('//div[contains(@class,"Specifications-descriptionList-cell")]')
                info_input_times = len(info_list) // 2
                for i in range(info_input_times):
                    key = info_list[0 + 2 * i].text
                    value = info_list[1 + 2 * i].text
                    if not self.infomation_dict.__contains__(key):
                        self.infomation_dict[key] = value
            time.sleep(3)
        # print(self.weight_dict)
        # print(self.infomation_dict)

    def get_sku(self,add=''):
        match='(?<=-)[a-z0-9]*(?=\.html)'
        sku=re.findall(match, add)
        if sku:
            return sku[0]
        return ''

    def get_cell(self, sheetnumber, row):
        '''获取表中网页地址，定义第几张表和列'''
        web_list=[]
        wb=openpyxl.load_workbook('wayfair.xlsx')
        sheetname_list=wb.sheetnames
        print("Totally have %s sheets."%len(sheetname_list))
        sheetname=sheetname_list[sheetnumber-1]
        # table=wb[sheetname]  #表名
        sheet = wb.get_sheet_by_name(sheetname)
        rows = sheet.max_row
        columns = sheet.max_column
        # cell=table[column]  #列
        print("column count is %s"%columns)
        self.current_type = sheet.cell(row=row, column=2).value
        for c in range(2, columns+1):
            if sheet.cell(row=row, column=c).value:
                if 'www.wayfair' in sheet.cell(row=row, column=c).value:
                    web_list.append(sheet.cell(row=row, column=c).value)
        print("row:%s get %s web links."%(row, len(web_list)))
        return web_list

    def write_cell(self, sheetnumber, columnnumber):
        '''写入信息至excel，定义第几张表，起始列'''
        wb = openpyxl.load_workbook('wayfair.xlsx')
        sheetname_list = wb.sheetnames
        sheetname = sheetname_list[sheetnumber - 1]
        sheet=wb.get_sheet_by_name(sheetname)
        self.logger.info("Start write to:%s, column number:%s"%(sheetname, columnnumber))
        sheet.cell(row=1, column=columnnumber).value = self.current_type
        sheet.merge_cells(start_row=1, end_row=1, start_column=columnnumber,
                          end_column=columnnumber + 1)
        # sheet.cell(row=2, column=columnnumber).value=self.weight_name
        # sheet.merge_cells(start_row=2, end_row=2, start_column=columnnumber,
        #                   end_column=columnnumber + 1)  # merge_cells('A2:D2')
        # print('Set title:%s' % self.weight_name)
        # i=3
        # for weight in self.weight_dict:
        #     sheet.cell(row=i, column=columnnumber).value=weight
        #     sheet.cell(row=i, column=columnnumber+1).value = self.weight_dict[weight]
        #     i+=1
        j=3
        sheet.cell(row=j, column=columnnumber).value = self.infomation_name
        sheet.merge_cells(start_row=j, end_row=j, start_column=columnnumber,
                          end_column=columnnumber + 1)
        print('Set title:%s' % self.infomation_name)
        for infomation in self.infomation_dict:
            sheet.cell(row=j+1, column=columnnumber).value=infomation
            sheet.cell(row=j+1, column=columnnumber+1).value= self.infomation_dict[infomation]
            j+=1
        wb.save('wayfair.xlsx')

    def test_wayfair(self, sheetnumber, column):
        self.set_up('https://www.wayfair.com')
        self.get_cell(sheetnumber, column)
        for add in self.web_list:
            sku=self.get_sku(add)
            if sku:
                self.search(sku)
                self.get_item_info()
        self.close()
        # print(self.weight_dict)
        # print(self.infomation_dict)
        self.write_cell(sheetnumber,3)

    def qingqing_wayfair(self):
        self.set_up('https://www.wayfair.co.uk')
        self.current_url = 'https://www.wayfair.co.uk'
        start_row = 2
        for i in range(start_row, 29):  #总共有58行
            self.logger.info("Start the %s row."%(start_row))
            web_list=self.get_cell(1, start_row)
            for add in web_list:
                sku = self.get_sku(add)
                if sku:
                    if self.search(sku):
                        self.get_item_info()
            # print(self.weight_dict)
            # print(self.infomation_dict)
            self.write_cell(2, -1+2*i)
            start_row += 1
            #一列执行完之后清空存储的数据组
            # self.weight_dict = {}  # 用来存储Weights & Dimensions相关信息
            self.infomation_dict = {}  # 用来存储Specifications相关信息
            # self.weight_name = ''  # 用来存储Weights & Dimensions的text的值
            self.infomation_name = ''  # 用来存储Specifications的text的值
            self.current_type = ''
        self.close()



if __name__ == '__main__':
    test = wayfair_test('wayfair_test')
    # test.total_test()
    # test.test_wayfair(1,1)
    test.qingqing_wayfair()
    # test.get_cell(1,'A')
    # test.write_cell(1,3)
    # test.get_sku('https://www.wayfair.com/furniture/pdp/gforce-black-tilt-wall-mount-holds-up-to-100-lbs-gfrc1085.html')

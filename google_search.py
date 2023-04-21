# -*- coding: utf-8 -*-
from selenium import webdriver
from requests_html import HTMLSession
from selenium.webdriver import ActionChains  # 处理鼠标悬停事件
from selenium.webdriver import TouchActions
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.common.by import By  # By类：定位元素
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.keys import Keys
import sys
import re
import os
import time, datetime
import random
import openpyxl
from openpyxl.drawing.image import Image
from root_scripts.common import Common
from PIL import Image as pilimage
from requests_toolbelt import MultipartEncoder  # 上传文件模块
from root_scripts.common import output_print

# Python3字符串默认编码unicode
libpath = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if not libpath in sys.path:
    sys.path.append(libpath)


class google_search_test(Common):
    # output_print=output_print()
    def __init__(self, logname):
        Common.__init__(self, logname)
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.212 Safari/537.36",
            "content-type": "application/x-www-form-urlencoded; charset=UTF-8"}  # "User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Safari/537.36",
        self.sessions = HTMLSession()
        self.sessions.headers = headers
        self.ad_items = []  # 用来存广告排名前10的商品信息
        self.natual_items = []  # 用来存自然流量排名前10的商品信息
        self.temp_item = {'name': '', 'price': '', 'source': '', 'address': ''}  # 临时存放数据的

    def set_up(self, web='https://www.google.com/search?q=table&tbm=shop', device_type='html'):
        device_type = device_type.upper()
        if device_type == 'H5':
            mobile_emulation = {'deviceName': 'iPhone X'}
            options = webdriver.ChromeOptions()
            options.add_experimental_option('mobileEmulation', mobile_emulation)
        else:
            options = webdriver.ChromeOptions()
        if os.path.exists("D:/python/chromedriver.exe"):
            self.browser = webdriver.Chrome("D:/python/chromedriver.exe", chrome_options=options)
        else:
            self.browser = webdriver.Chrome("C:/Users/aosom/AppData/Local/Google/Chrome/Application/chromedriver.exe",
                                            chrome_options=options)
        try:
            self.browser.implicitly_wait(10)  # 隐式等待10秒，如果在10秒内网页加载完成，则执行下一步，否则一直等到时间截止，然后执行下一步
            self.browser.get(web)  # 在当前窗口加载 url
            self.logger.info("Start web:%s" % web)
            self.browser.maximize_window()
            print("Set window to max size.")
        except TimeoutException:
            print("Can not open the web!")

    def close(self):
        self.browser.close()  # 关闭当前窗口, 如果当前窗口是最后一个窗口, 浏览器将关闭
        self.browser.quit()  # 关闭所有窗口并停止 ChromeDriver 的执行
        print("Close the browser.")

    def skip_search_info_page(self):
        try:
            buttons = self.browser.find_elements_by_xpath('//div[@class="VfPpkd-dgl2Hf-ppHlrf-sM5MNb"]/button')
            if buttons:
                buttons[1].click()  # 点第二个我同意跳过
                time.sleep(3)
        except:
            return True

    def get_page_info(self):
        """获取整个页面内容"""
        result = self.sessions.get(
            "https://www.google.com/search?q=site:www.aosom.com+inurl:/keyword&ei=pIlfY5H7C5Go5NoP_uqnkAQ&start=10&sa=N&ved=2ahUKEwiR3fbPiIr7AhURFFkFHX71CUIQ8tMDegQIHBAE&biw=1326&bih=1033&dpr=0.9")
        print(result.text)

    def search(self, keyword=''):
        self.browser.find_element_by_class_name('gLFyf').clear()
        self.browser.find_element_by_class_name('gLFyf').send_keys(keyword)
        self.browser.find_element_by_class_name('gLFyf').send_keys(Keys.ENTER)
        print("Search keyword:%s" % keyword)
        time.sleep(10)  # 搜索完多等待一会

    def get_redirects_add(self, add=''):
        '''获得302跳转页面的最终页面链接'''
        # if add.split(':')[0] == 'https':
        #     result = self.sessions.get(url=add, proxies='https')
        # else:
        #     result = self.sessions.get(url=add, proxies='http')
        result = self.sessions.get(url=add)
        his_add = result.history
        # print (his_add)
        return his_add[-1].headers["location"]

    def deal_name(self, name=''):
        '''处理字符串，去掉中间特殊的字符和空格'''
        deal_name = "".join(filter(str.isalnum, name))
        # print(deal_name)
        return deal_name

    def get_page_adinfo(self, word_count=10):
        '''获取页面ad部分关键信息,根据关键字的个数来控制抓的条数'''
        self.ad_items = []  # 每一轮执行前都清除上一轮的数组
        aditems = self.browser.find_elements_by_class_name('KZmu8e')
        total_count = len(aditems)
        if word_count > 30:
            if total_count > 10:
                loop_count = 10
            else:
                loop_count = total_count
        elif 15 <= word_count <= 30:
            if total_count > 20:
                loop_count = 20
            else:
                loop_count = total_count
        else:
            if total_count > 50:
                loop_count = 50
            else:
                loop_count = total_count
        print("Totally find %s matched ad items,Get %s items info." % (total_count, loop_count))
        for i in range(loop_count):
            temp_item = {'name': '', 'price': '', 'source': '', 'address': '', 'photo_add': ''}
            if self.iselementexits_by_classname(self.browser, 'sh-np__product-title'):
                name = aditems[i].find_element_by_class_name('sh-np__product-title').text  # 可能class多加了一个内容
            else:
                name = aditems[i].find_element_by_class_name('translate-content').text
            # name=aditems[i].find_element_by_xpath('//div[contains(@class,"product-title")]').text
            # name_list=aditems[i].find_elements_by_class_name('sh-np__product-title')
            # print('共找到了%s个名字参数'%len(name_list))
            if self.iselementexits_by_classname(self.browser, 'T14wmb'):
                price = aditems[i].find_element_by_class_name('T14wmb').text
            else:
                price = ''
            if self.iselementexits_by_classname(self.browser, 'E5ocAb'):
                source = aditems[i].find_element_by_class_name('E5ocAb').text
            else:
                source = ''
            # google_add=aditems[i].find_element_by_class_name('sh-np__click-target').get_attribute('href')  #'https://www.google.com'+
            if self.iselementexits_by_classname(self.browser, 'sh-np__click-target'):
                google_add = aditems[i].find_element_by_class_name('sh-np__click-target').get_attribute(
                    'href')  # shntl sh-np-h__click-target
                # photo_add= aditems[i].find_element_by_xpath('//div[@class="SirUVb sh-img__image"]/img[1]').get_attribute('src')
                photo_add_path = aditems[i].find_element_by_class_name('sh-np__click-target').get_attribute('innerHTML')
            # <div class="SirUVb sh-img__image" style="padding:0px;width:168px;height:168px">
            # <img src="https://encrypted-tbn3.gstatic.com/shopping?q=tbn:ANd9GcSXtzZULIv6jmoiJ5jtP0cMztl-ZHvM2dfhV9kLp8ajpzY7QwFOt-3AGToiWyjVClkqfuGMfX4-vRoNPiLiz-UJf1IdK4lIbS1_025RYoXY&amp;usqp=CAE"
            # alt="" role="presentation" data-atf="0"></div><div class="rz2LD"></div><div class="HUOptb">
            # <div class="ljqwrc"><div class="sh-np__product-title translate-content" data-merchant-id="130826929" data-offer-id="18177">
            # Outflexx Gewächshaus anthrazit Alu/Polycarbonat 271x271x241/180cm 4 Fenster 10 mm Wandstärke</div><div class="hn9kf"><span class="T14wmb">
            # <b>€2,179.00</b></span></div><div class="sh-np__seller-container" data-merchant-id="130826929" data-offer-id="18177" aria-label="来自Gartenmoebel.de" role="link">
            # <span class="E5ocAb">Gartenmoebel.de</span></div><div class="U6puSd">免费送货</div></div></div>
            else:
                google_add = aditems[i].find_element_by_class_name('shntl').get_attribute('href')
                # photo_add =aditems[i].find_element_by_xpath('//div[@class="SirUVb sh-img__image"]/img[1]').get_attribute('src')
                # google_add = aditems[i].find_element_by_xpath('//a[contains(@class,"click-target")]').get_attribute('href')
                photo_add_path = aditems[i].find_element_by_class_name('shntl').get_attribute('innerHTML')
            # print(type(photo_add_path))
            # print(photo_add_path)
            # match = "(?<=src=(\"|\')).*?(?=\"|\')"
            match = "(?<=src=\")https.*?(?=\")"
            photo_add = re.findall(match, photo_add_path)
            if len(photo_add) == 0:
                photo_add = ''
            else:
                photo_add = photo_add[0]
            # print(photo_add_path)
            # photo_add=aditems[i].find_element_by_xpath('//div[contains(@class,"img__image")]/img').get_attribute('src')
            # if self.iselementexits_by_xpath(self.browser, '//div[@class="sh-img__image"]/img'):
            #     photo_add = aditems[i].find_element_by_xpath('//div[@class="sh-img__image"]/img').get_attribute('src')
            # else:
            #     photo_add = aditems[i].find_element_by_xpath('//div[@class="SirUVb"]/img').get_attribute('src')
            print('loop %s photo add is :%s' % (i + 1, photo_add))
            # save_name=self.deal_name(name)
            # self.download_image(photo_add, save_name)
            # try:
            #     add=self.get_redirects_add(google_add)
            # except:
            #     add=google_add
            if photo_add:
                save_name = self.deal_name(name)
                self.download_image(photo_add, save_name)
                try:
                    add = self.get_redirects_add(google_add)
                except:
                    add = google_add
            else:
                add = google_add
            temp_item['name'] = name
            temp_item['price'] = price
            temp_item['source'] = source
            temp_item['address'] = add
            temp_item['photo_add'] = photo_add
            self.ad_items.append(temp_item)

    def get_page_natualinfo(self, word_count=10):
        '''获取页面natual部分关键信息,根据关键字的个数来控制抓的条数'''
        self.natual_items = []  # 每一轮执行前都清除上一轮的数组
        aditems = self.browser.find_elements_by_class_name('sh-dgr__grid-result')
        total_count = len(aditems)
        if word_count > 30:
            if total_count > 10:
                loop_count = 10
            else:
                loop_count = total_count
        elif 15 <= word_count <= 30:
            if total_count > 20:
                loop_count = 20
            else:
                loop_count = total_count
        else:
            if total_count > 50:
                loop_count = 50
            else:
                loop_count = total_count
        print("Totally find %s matched natual items,Get %s items info." % (total_count, loop_count))
        for i in range(loop_count):
            temp_item = {'name': '', 'price': '', 'source': '', 'address': '', 'photo_add': ''}
            if self.iselementexits_by_classname(self.browser, 'A2sOrd'):
                name = aditems[i].find_element_by_class_name('A2sOrd').text
            elif self.iselementexits_by_classname(self.browser, 'Xjkr3b'):
                name = aditems[i].find_element_by_class_name('Xjkr3b').text
            else:
                name = ''
            if self.iselementexits_by_classname(self.browser, 'a8Pemb'):
                price = aditems[i].find_element_by_class_name('a8Pemb').text
            else:
                price = ''
            if self.iselementexits_by_classname(self.browser, 'aULzUe'):
                source = aditems[i].find_element_by_class_name('aULzUe').text
            else:
                source = ''
            google_add = aditems[i].find_element_by_class_name('xCpuod').get_attribute(
                'href')  # 'https://www.google.com'+
            # photo_add = aditems[i].find_element_by_class_name('//div[@class="ArOc1c"]/img').get_attribute('src')
            # photo_add = aditems[i].find_element_by_xpath('//div[@class="ArOc1c"]/img[1]').get_attribute('src')
            try:
                aditems[i].find_element_by_class_name("ArOc1c").click()  # 先点击展开商品的扩展框
                time.sleep(4)
            except:
                pass
            if "https://www.google.com" in google_add:
                google_add_part = google_add.replace('https://www.google.com', '')
            else:
                google_add_part = google_add
            photo_xpath = '//a[@href="' + google_add_part + '"]/div[1]/div[1]'
            if self.iselementexits_by_xpath(self.browser, photo_xpath):
                photo_add_path = aditems[i].find_element_by_xpath(photo_xpath).get_attribute(
                    'innerHTML')  # 获取元素内的全部HTML
                # print(photo_add_path)
                if photo_add_path:
                    match = "(?<=src=\")https.*?(?=\")"
                    photo_add = re.findall(match, photo_add_path)
                else:
                    photo_add = ''
            else:
                photo_add = ''
            if len(photo_add) == 0:
                photo_add = ''
            else:
                photo_add = photo_add[0]
            print('loop %s photo add is :%s' % (i + 1, photo_add))
            if photo_add:
                save_name = self.deal_name(name)
                self.download_image(photo_add, save_name)
                try:
                    add = self.get_redirects_add(google_add)
                except:
                    add = google_add
            else:
                add = google_add
            temp_item['name'] = name
            temp_item['price'] = price
            temp_item['source'] = source
            temp_item['address'] = add
            temp_item['photo_add'] = photo_add
            self.natual_items.append(temp_item)

    def create_excel(self, country='CN'):
        """根据执行时间新建excel,表名中带国家名字"""
        excel_new = openpyxl.Workbook()
        sheet1 = excel_new.active
        sheet1.title = "google_keyword_result"
        t = time.strftime("%Y_%m_%d", time.localtime())  # _%H_%M
        if country:
            filename = 'results/' + t + '_' + country + '_google_keyword.xlsx'
        else:
            filename = 'results/' + t + '_google_keyword.xlsx'
        excel_new.save(filename)
        print("Create file:%s" % filename)
        return filename

    def create_excel_sheet(self, filename='', sheetname=''):
        """新建表格的sheet"""
        if not os.path.exists(filename):
            excel_new = openpyxl.Workbook()
            sheet1 = excel_new.active
            sheet1.title = sheetname
            excel_new.save(filename)
            print("Create file:%s, and create new sheet:%s" % (filename,sheetname))
            return filename
        wb = openpyxl.load_workbook(filename)
        sheetname_list = wb.sheetnames
        if sheetname in sheetname_list:
            print('Already have sheet name:%s' % sheetname)
            return filename
        # sheet_amount= len(sheetname_list)
        wb.create_sheet(sheetname)
        # sheet.title = sheetname
        wb.save(filename)
        print("Create new sheet:%s" % sheetname)
        return filename

    def write_excel(self, filename='', keyword=''):
        """将数组中存放的内容写入excel,需求纵向排序在下方追加写入"""
        wb = openpyxl.load_workbook(filename)
        sheetname_list = wb.sheetnames
        sheetname = sheetname_list[0]
        sheet = wb.get_sheet_by_name(sheetname)
        # columns = sheet.max_column  #横向扩展，获取最大的列数
        # if columns==1:
        #     columns=0
        columns = 0
        rows = sheet.max_row
        if rows == 1:
            rows = 0
        sheet.cell(row=1 + rows, column=columns + 1).value = keyword
        sheet.cell(row=2 + rows, column=columns + 1).value = '商品名'
        sheet.cell(row=2 + rows, column=columns + 2).value = '商品图片'
        sheet.cell(row=2 + rows, column=columns + 3).value = '商品价格'
        sheet.cell(row=2 + rows, column=columns + 4).value = '商品来源网站'
        sheet.cell(row=2 + rows, column=columns + 5).value = '商品链接'
        sheet.cell(row=3 + rows, column=columns + 1).value = "广告流量排行"
        if self.ad_items:
            sheet.column_dimensions['B'].width = 20  # 列宽
            print("Input ad items into excel")
            for i in range(len(self.ad_items)):  # 循环输入数组中的字典
                sheet.cell(row=4 + rows + i, column=columns + 1).value = self.ad_items[i]['name']
                photo_name = self.deal_name(self.ad_items[i]['name'])
                image_path = 'results/' + photo_name + '.jpg'
                if os.path.exists(image_path):
                    image = pilimage.open(image_path)
                    try:
                        image.save(image_path, 'JPEG')
                        img = Image(image_path)  # 选择你的图片
                        img.width, img.height = 75, 75
                        sheet.add_image(img, 'B' + str(4 + rows + i))
                        sheet.row_dimensions[4 + rows + i].height = 70  # 行高
                    except Exception as e:
                        print(e)
                    # os.remove(image_path)
                else:
                    sheet.cell(row=4 + rows + i, column=columns + 2).value = self.ad_items[i]['photo_add']
                sheet.cell(row=4 + rows + i, column=columns + 3).value = self.ad_items[i]['price']
                sheet.cell(row=4 + rows + i, column=columns + 4).value = self.ad_items[i]['source']
                sheet.cell(row=4 + rows + i, column=columns + 5).value = self.ad_items[i]['address']
        sheet.cell(row=4 + len(self.ad_items) + rows, column=columns + 1).value = "自然流量排行"
        if self.natual_items:
            print("Input natual items into excel")
            for i in range(len(self.natual_items)):  # 循环输入数组中的字典
                sheet.cell(row=5 + len(self.ad_items) + rows + i, column=columns + 1).value = self.natual_items[i][
                    'name']
                photo_name = self.deal_name(self.natual_items[i]['name'])
                image_path = 'results/' + photo_name + '.jpg'
                if os.path.exists(image_path):
                    image = pilimage.open(image_path)
                    image.save(image_path, 'JPEG')
                    img = Image(image_path)  # 选择你的图片
                    img.width, img.height = 75, 75
                    sheet.add_image(img, 'B' + str(5 + len(self.ad_items) + rows + i))
                    sheet.row_dimensions[5 + len(self.ad_items) + rows + i].height = 70  # 行高
                    # os.remove(image_path)
                else:
                    sheet.cell(row=5 + len(self.ad_items) + rows + i, column=columns + 2).value = self.natual_items[i][
                        'photo_add']
                sheet.cell(row=5 + len(self.ad_items) + rows + i, column=columns + 3).value = self.natual_items[i][
                    'price']
                sheet.cell(row=5 + len(self.ad_items) + rows + i, column=columns + 4).value = self.natual_items[i][
                    'source']
                sheet.cell(row=5 + len(self.ad_items) + rows + i, column=columns + 5).value = self.natual_items[i][
                    'address']
        wb.save(filename)

    def download_image(self, add='', name=''):
        '''下载图片'''
        if add == '':
            print("Address is empty, skip!")
            return False
        try:
            res = self.sessions.get(add)
            file_name = 'results/' + name + '.jpg'
            with open(file_name, 'wb') as f:
                f.write(res.content)
                print("Download photo %s.jpg succeed!" % name)
            f.close()
        except:
            print("Download photo %s.jpg failed!" % name)
            return False

    def write_image(self, filename='results/2021_07_17_16_12_google_keyword.xlsx', image_name='ad.jpg'):
        '''把图片添加进excel'''
        wb = openpyxl.load_workbook(filename)
        sheetname_list = wb.sheetnames
        sheetname = sheetname_list[0]
        sheet = wb.get_sheet_by_name(sheetname)
        img = Image('results/' + image_name)  # 选择你的图片
        img.width, img.height = 80, 80
        sheet.add_image(img, 'B20')
        sheet.column_dimensions['B'].width = 20  # 列宽
        sheet.row_dimensions[20].height = 70  # 行高
        wb.save(filename)

    def get_country_list(self):
        '''读取excel表中国家列表'''
        wb = openpyxl.load_workbook('谷歌抓取主推类目.xlsx')
        sheetname_list = wb.sheetnames
        # print(sheetname_list)
        return sheetname_list

    def select_country(self, country_list=[]):
        '''从国家列表选择国家'''
        '''将国家列表展示出来，并选择'''
        for i in range(len(country_list)):
            print("%s:%s" % (i + 1, country_list[i]))
        country_num = input("PLS select the country ID:")
        country_num = int(country_num)
        if 0 < country_num <= len(country_list):
            print("You have select:%s" % country_list[country_num - 1])
            return country_list[country_num - 1]
        else:
            print("Country ID error!")
            self.select_country(country_list)

    def get_keyword_list(self, country_name=''):
        wb = openpyxl.load_workbook('谷歌抓取主推类目.xlsx')
        sheet = wb.get_sheet_by_name(country_name)
        row = sheet.max_row
        keyword_list = []
        for i in range(row):
            if sheet.cell(row=1 + i, column=1).value:
                keyword_list.append(sheet.cell(row=1 + i, column=1).value)
        print(keyword_list)
        return keyword_list

    def upload_file(self, filename='', filepath=''):
        '''上传文件'''
        headers = {
            'Content-Type': 'multipart/form-data; boundary=----WebKitFormBoundaryHF2C97Tky0OzQDQm',
        }
        # data = {
        #     'name': filename,
        #     # 'Content-Disposition': 'form-data',
        #     # 'Content-Type': 'multipart/form-data',
        #     # 'boundary':'',
        #     # "Content-Length":'',
        #     # 'Content-Type': 'application/vnd.ms-excel'
        # }
        data = MultipartEncoder(fields={'file': (filename, open(filepath, 'rb'), 'application/vnd.ms-excel'), }
                                # 'identityType':identityType
                                , boundary='----WebKitFormBoundaryHF2C97Tky0OzQDQm')
        # with open(filepath, 'rb') as files:
        # files = {'file': open(filepath, 'rb')}
        # files = {'file': (filename, open(filepath, 'rb'), 'application/vnd.ms-excel', {'Expires': '0'})}
        print('Upload file %s' % filename)
        response = self.sessions.post('https://www.aosom.co.uk/aosom_ca/product/infoupload', headers=headers,
                                      data=data)  # files=files
        print(response.text)
        return response

    def check_time(self, set_time='22'):
        '''判断当前时间,当到达匹配时间返回True'''
        t = time.strftime("%H", time.localtime())
        print("It's %s o'clock." % t)
        if t == set_time:
            return True
        return False

    def delete_photo(self):
        '''清空爬下来的所有图片'''
        files = os.listdir('results/')
        for file in files:
            if '.jpg' in file:
                os.remove('results/' + file)
        # print(files)

    def write_search_result_to_excel(self, filename='', sheetname='', title='', url=''):
        '''将数组中存放的内容写入excel,需求纵向排序在下方追加写入'''
        wb = openpyxl.load_workbook(filename)
        # sheetname_list = wb.sheetnames
        # sheetname = sheetname_list[0]
        sheet = wb.get_sheet_by_name(sheetname)
        # columns = sheet.max_column  #横向扩展，获取最大的列数
        # if columns==1:
        #     columns=0
        columns = 0
        rows = sheet.max_row
        # if rows == 1:
        #     rows = 0
        sheet.cell(row=rows + 1, column=1).value = title
        sheet.cell(row=rows + 1, column=2).value = url
        wb.save(filename)

    def keyword_search_test(self):
        """搜索关键词，获取url和内容"""
        self.set_up()
        filename = self.create_excel('')
        url_list = ['CA', 'DE', 'ES', 'FR', 'IT', 'PT', 'co.UK', 'RO', 'IE', 'com']  #
        for current_url in url_list:
            current_url = current_url.lower()
            self.create_excel_sheet(filename, current_url)
            loop_time = 0
            for i in range(30):
                num = str(loop_time * 10)
                url = "https://www.google." + current_url + "/search?q=site:www.aosom." + current_url + \
                      "+inurl:/keyword&ei=sz1rY5OuD6LW5NoPlKC64Ao&start=" + num + "&sa=N&" \
                      "ved=2ahUKEwjci4ewhYr7AhX2L1kFHWSLD3M4MhDy0wN6BAgdEAQ&biw=1326&bih=1033&dpr=0.9"
                delay = random.randint(1, 5)
                time.sleep(delay)
                self.browser.get(url)
                url_list = self.browser.find_elements_by_xpath("//div[@class='yuRUbf']/a")
                if url_list:
                    for u in url_list:
                        siteurl = u.get_attribute('href')
                        sitetitle = u.text
                        print(sitetitle, siteurl)
                        self.write_search_result_to_excel(filename, current_url, sitetitle, siteurl)
                    loop_time += 1
                # elif self.browser.find_element_by_id("recaptcha").is_displayed():  #人机验证
                elif self.iselementexist_by_id(self.browser, "recaptcha"):
                    while True:
                        print("have google robot, delay 15S")
                        time.sleep(15)
                        if self.iselementexist_by_id(self.browser, "recaptcha"):
                            continue
                        # url_list = self.browser.find_elements_by_xpath("//div[@class='yuRUbf']/a")
                        # if not url_list:
                        #     print("have google robot, delay 15S")
                        #     time.sleep(15)
                        #     continue
                        # loop_time -= 1
                        break
                else:
                    break

    def total_test(self):
        country_list = self.get_country_list()
        country = self.select_country(country_list)
        keyword_list = self.get_keyword_list(country)
        keyword_count = len(keyword_list)
        while True:
            if self.check_time('22'):
                filename = self.create_excel(country)
                # self.set_up('https://www.google.com/search?q=hollywoodschaukeln&tbm=shop&ei=HnPyYOfTIeGDlQfHwZSQBA&oq=hollywoodschaukeln&gs_lcp=Cgtwcm9kdWN0cy1jYxADULEqWNgqYOsqaABwAHgAgAEAiAEAkgEAmAEAoAEBwAEB&sclient=products-cc&ved=0ahUKEwjn2tPot-nxAhXhQeUKHccgBUIQ4dUDCAs&uact=5')
                self.set_up()
                self.skip_search_info_page()
                for keyword in keyword_list:
                    self.search(keyword)
                    self.get_page_adinfo(keyword_count)
                    self.get_page_natualinfo(keyword_count)
                    self.write_excel(filename, keyword)
                self.close()
                filename_short = filename.replace('results/', '')
                self.upload_file(filename_short, filename)
                self.delete_photo()
                time.sleep(72000)  # 完成后延迟20小时
            time.sleep(600)


if __name__ == '__main__':
    test = google_search_test('test')
    # sys.stdout = output_print
    # test.delete_photo()
    test.keyword_search_test()
    # test.total_test()
    # test.download_image('https://encrypted-tbn0.gstatic.com/shopping?q=tbn:ANd9GcRRAPnyZRluoDp05SlDyTfTsnl2guIQiBPTF2OGCJRWc71bSN_cVsqBKupq-3UP8n6CLS-A06xAVGTpAlfCEjaxP7-KsyXPKcqQWBmuvCfd&usqp=CAE','ttt.jpg')
    # test.get_country_list()
